# ==========================================================
# Store Dashboard — Cart-based Ordering System
# ==========================================================

import io
import re
from datetime import datetime
import pytz
import streamlit as st
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from pathlib import Path

st.set_page_config(page_title="Store Dashboard", layout="wide")

# ----------------------------------------------------------
# CSS
# ----------------------------------------------------------
def load_css():
    css_file = Path(__file__).parent / "styles.css"
    if css_file.exists():
        with open(css_file) as f:
            st.markdown(f"<style>{f.read()}</style>", unsafe_allow_html=True)

load_css()
st.title("🏪 Store Dashboard")

# ==========================================================
# SUPABASE CONNECTION
# ==========================================================
def get_secret(key, default=None):
    try:
        return st.secrets[key]
    except Exception:
        return default

SUPABASE_URL = get_secret("SUPABASE_URL")
SUPABASE_KEY = get_secret("SUPABASE_KEY")

SUPABASE_ENABLED = SUPABASE_URL is not None and SUPABASE_KEY is not None

if not SUPABASE_ENABLED:
    st.error(
        "❌ Supabase secrets not found. "
        "Add **SUPABASE_URL** and **SUPABASE_KEY** to your Streamlit secrets."
    )
    st.stop()

@st.cache_resource
def get_supabase_client():
    from supabase import create_client
    return create_client(SUPABASE_URL, SUPABASE_KEY)

# ==========================================================
# PAGINATED FETCH
# ==========================================================
def fetch_all(table: str, columns: str = "*", filters: dict = None) -> list[dict]:
    client   = get_supabase_client()
    page     = 0
    size     = 1000
    all_rows = []

    while True:
        query = client.table(table).select(columns).range(page * size, (page + 1) * size - 1)
        if filters:
            for col, val in filters.items():
                query = query.eq(col, val)
        result = query.execute()
        batch  = result.data or []
        all_rows.extend(batch)
        if len(batch) < size:
            break
        page += 1

    return all_rows

# ==========================================================
# CART FUNCTIONS
# ==========================================================
def add_to_cart(items_df):
    """Add items to cart (upsert based on plu_code)"""
    client = get_supabase_client()

    records = []
    for _, row in items_df.iterrows():
        records.append({
            "plu_code": str(row["PLU CODE"]),
            "description": str(row.get("DESCRIPTION", "")),
            "cost_price": float(row.get("COST PRICE", 0) or 0),
            "selling_price": float(row.get("SELLING PRICE", 0) or 0),
            "supplier": str(row.get("SUPPLIER", "")),
            "category": str(row.get("CATEGORY", "")),
            "group_info": str(row.get("GROUP", "")),
            "stock": int(pd.to_numeric(row.get("STOCK", 0), errors="coerce") or 0),
            "usage": int(pd.to_numeric(row.get("USAGE", 0), errors="coerce") or 0),
            "order_qty": int(pd.to_numeric(row["ORDER QTY"], errors="coerce") or 0),
            "added_at": datetime.now(pytz.UTC).isoformat()
        })

    for record in records:
        client.table("cart_items").upsert(record, on_conflict="plu_code").execute()

    return len(records)

def get_cart():
    """Fetch all cart items"""
    rows = fetch_all("cart_items")
    if not rows:
        return pd.DataFrame()

    df = pd.DataFrame(rows)
    df = df.rename(columns={
        "plu_code": "PLU CODE",
        "description": "DESCRIPTION",
        "cost_price": "COST PRICE",
        "selling_price": "SELLING PRICE",
        "supplier": "SUPPLIER",
        "category": "CATEGORY",
        "group_info": "GROUP",
        "stock": "STOCK",
        "usage": "USAGE",
        "order_qty": "ORDER QTY",
        "added_at": "ADDED AT"
    })
    return df

def get_cart_count():
    """Get total items in cart"""
    client = get_supabase_client()
    result = client.table("cart_items").select("plu_code", count="exact").execute()
    return result.count or 0

def clear_cart():
    """Clear entire cart"""
    client = get_supabase_client()
    client.table("cart_items").delete().neq("plu_code", "").execute()

def clear_supplier_cart(supplier):
    """Clear items for specific supplier"""
    client = get_supabase_client()
    client.table("cart_items").delete().eq("supplier", supplier).execute()

def delete_cart_item(plu_code):
    """Delete single item from cart"""
    client = get_supabase_client()
    client.table("cart_items").delete().eq("plu_code", plu_code).execute()

# ==========================================================
# DATA LOADERS
# ==========================================================
@st.cache_data(ttl=300)
def load_prices() -> pd.DataFrame:
    rows = fetch_all("existing_prices")
    df   = pd.DataFrame(rows)
    if df.empty:
        return df

    df = df.rename(columns={
        "barcode":     "BARCODE",
        "item_num":    "ITEM NUM",
        "description": "Description",
        "size":        "Size",
        "pack":        "Pack",
        "price":       "Price",
        "pc_cost":     "Pc. Cost",
        "sell_price":  "Sell Price",
        "aisle":       "AISLE",
        "supplier":    "SUPPLIER",
    })

    df = df[[c for c in [
        "BARCODE","ITEM NUM","Description","Size","Pack",
        "Price","Pc. Cost","Sell Price","AISLE","SUPPLIER"
    ] if c in df.columns]]

    for col in ["Price", "Pc. Cost", "Sell Price"]:
        if col in df.columns:
            df[col] = pd.to_numeric(df[col], errors="coerce")

    return df


@st.cache_data(ttl=300)
def load_yellow_basic() -> pd.DataFrame:
    rows = fetch_all("re_order", "plu_code, stock, usage", {"row_color": "yellow"})
    df   = pd.DataFrame(rows)
    if df.empty:
        return pd.DataFrame(columns=["PLU CODE","STOCK","USAGE"])

    df = df.rename(columns={"plu_code":"PLU CODE","stock":"STOCK","usage":"USAGE"})
    df = df.drop_duplicates(subset=["PLU CODE"], keep="first")

    st.success(f"✅ Loaded {df['PLU CODE'].nunique()} yellow PLU items")
    return df


@st.cache_data(ttl=300)
def load_yellow_full() -> pd.DataFrame:
    rows = fetch_all(
        "re_order",
        "plu_code, description, cost, group_info, stock, usage, supplier",
        {"row_color": "yellow"}
    )
    df = pd.DataFrame(rows)
    if df.empty:
        return pd.DataFrame(columns=["PLU CODE","DESCRIPTION","COST","GROUP","GROUP2","STOCK","USAGE"])

    df = df.rename(columns={
        "plu_code":    "PLU CODE",
        "description": "DESCRIPTION",
        "cost":        "COST",
        "group_info":  "GROUP",
        "stock":       "STOCK",
        "usage":       "USAGE",
    })

    df = df.drop_duplicates(subset=["PLU CODE"], keep="first")

    def clean_group2(val):
        g2 = str(val).strip() if val is not None else ""
        return "" if (g2 == "" or g2 == "0" or g2.lower() == "none") else g2

    df["GROUP2"] = df["supplier"].apply(clean_group2)
    return df[["PLU CODE","DESCRIPTION","COST","GROUP","GROUP2","STOCK","USAGE"]]


@st.cache_data(ttl=300)
def load_unordered() -> pd.DataFrame:
    rows = fetch_all(
        "re_order",
        "plu_code, description, cost, group_info, stock, price_1, usage, supplier",
        {"row_color": "none"}
    )
    df = pd.DataFrame(rows)
    if df.empty:
        return pd.DataFrame(columns=["PLU CODE","DESCRIPTION","COST PRICE","SELLING PRICE","GROUP","STOCK","USAGE","SUPPLIER"])

    df = df.rename(columns={
        "plu_code":    "PLU CODE",
        "description": "DESCRIPTION",
        "cost":        "COST PRICE",
        "group_info":  "GROUP",
        "stock":       "STOCK",
        "price_1":     "SELLING PRICE",
        "usage":       "USAGE",
        "supplier":    "SUPPLIER",
    })
    return df[["PLU CODE","DESCRIPTION","COST PRICE","SELLING PRICE","GROUP","STOCK","USAGE","SUPPLIER"]]


@st.cache_data(ttl=300)
def load_reorder_price1() -> pd.DataFrame:
    rows = fetch_all("re_order", "plu_code, price_1", {"row_color": "yellow"})
    df   = pd.DataFrame(rows)
    if df.empty:
        return pd.DataFrame(columns=["PLU CODE","PRICE 1"])

    df = df.rename(columns={"plu_code":"PLU CODE","price_1":"PRICE 1"})
    df = df.drop_duplicates(subset=["PLU CODE"], keep="first")
    return df

# ==========================================================
# REFRESH CONTROL
# ==========================================================
st.markdown("---")
col_refresh, col_ts, _ = st.columns([1.5, 3, 5])

with col_refresh:
    if st.button("🔄 Refresh Data", use_container_width=True):
        load_prices.clear()
        load_yellow_basic.clear()
        load_yellow_full.clear()
        load_unordered.clear()
        load_reorder_price1.clear()
        st.rerun()

cst     = pytz.timezone("America/Chicago")
now_cst = datetime.now(cst).strftime("%Y-%m-%d %I:%M %p CST")
with col_ts:
    st.caption(f"🕐 Data loaded at {now_cst} — auto-refreshes every 5 minutes")

st.markdown("---")

# ==========================================================
# GROUP HELPERS
# ==========================================================
def clean_group(val):
    if val is None:
        return ""
    return str(val).replace("\r\n","").replace("\r","").replace("\n","").replace("_x000D_","").strip()

def get_category(g):
    parts = re.findall(r"\[([^\]]+)\]", str(g))
    return parts[0].strip() if parts else ""

def get_suppliers(g):
    parts = re.findall(r"\[([^\]]+)\]", str(g))
    return [p.strip() for p in parts[1:] if p.strip()]

# ==========================================================
# EXCEL ORDER SHEET BUILDER
# ==========================================================
def build_order_excel(df_items, filename_prefix="order"):
    """Build Excel order sheet from items DataFrame"""
    wb = Workbook()
    ws = wb.active
    ws.title = "Order Sheet"

    hf    = Font(name="Arial", bold=True, color="FFFFFF", size=11)
    hfill = PatternFill("solid", fgColor="1F4E79")
    ha    = Alignment(horizontal="center", vertical="center", wrap_text=True)
    la    = Alignment(horizontal="left",   vertical="center")
    ra    = Alignment(horizontal="right",  vertical="center")
    ca    = Alignment(horizontal="center", vertical="center")
    qfill = PatternFill("solid", fgColor="E2EFDA")
    qfont = Font(name="Arial", bold=True, size=11)
    bdr   = Border(left=Side(style="thin"), right=Side(style="thin"),
                   top=Side(style="thin"),  bottom=Side(style="thin"))

    cols   = ["PLU CODE","DESCRIPTION","COST PRICE","SELLING PRICE","GROUP","STOCK","USAGE","ORDER QTY"]
    widths = [18, 35, 13, 14, 38, 10, 10, 13]

    ws.row_dimensions[1].height = 30
    for ci, (cn, w) in enumerate(zip(cols, widths), 1):
        c = ws.cell(1, ci, cn)
        c.font = hf
        c.fill = hfill
        c.alignment = ha
        c.border = bdr
        ws.column_dimensions[c.column_letter].width = w

    alt = PatternFill("solid", fgColor="F5F5F5")
    for ri, (_, row) in enumerate(df_items.iterrows(), 2):
        ws.row_dimensions[ri].height = 18
        bg = alt if ri % 2 == 0 else None
        for ci, cn in enumerate(cols, 1):
            c = ws.cell(ri, ci, row.get(cn, ""))
            c.border = bdr
            c.font = Font(name="Arial", size=10)

            if cn in ("COST PRICE","SELLING PRICE","STOCK","USAGE","ORDER QTY"):
                c.alignment = ra
            elif cn == "PLU CODE":
                c.alignment = ca
            else:
                c.alignment = la

            if cn == "ORDER QTY":
                c.fill = qfill
                c.font = qfont
            elif bg:
                c.fill = bg

    ws.freeze_panes = "A2"
    sr = len(df_items) + 2
    ws.cell(sr, 1, "TOTAL ITEMS").font = Font(name="Arial", bold=True, size=11)
    ws.cell(sr, 1).alignment = la
    ws.cell(sr, 8, f"=SUM(H2:H{sr-1})").font = Font(name="Arial", bold=True, size=11)
    ws.cell(sr, 8).alignment = ra
    ws.cell(sr, 8).fill = qfill

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()

# ==========================================================
# LOAD ALL DATA
# ==========================================================
df_prices    = load_prices()
df_ybasic    = load_yellow_basic()
df_yfull     = load_yellow_full()
df_unordered = load_unordered()
df_price1    = load_reorder_price1()

if df_prices.empty:
    st.error("❌ No data in existing_prices table. Run the file watcher to sync your Excel files.")
    st.stop()

if not df_ybasic.empty:
    df_search = df_prices.merge(df_ybasic, left_on="BARCODE", right_on="PLU CODE", how="left")
    df_search = df_search.drop(columns=["PLU CODE"], errors="ignore")
else:
    df_search = df_prices.copy()
    df_search["STOCK"] = ""
    df_search["USAGE"] = ""

reorder_available = not df_yfull.empty or not df_unordered.empty

# ==========================================================
# SESSION STATE
# ==========================================================
for k, v in [
    ("order_clear", 0),
    ("search_clear", 0),
    ("sv_clear", 0),
    ("sv_mode", "Category"),
    ("last_search", ""),
    ("active_tab", "📋 Browse & Add to Cart"),
    ("cart_refresh", 0),
]:
    if k not in st.session_state:
        st.session_state[k] = v

# Persist draft quantities in Tab 1 across reruns/tab switches
if "draft_order_qty" not in st.session_state:
    st.session_state.draft_order_qty = {}

# Get cart count for badge
cart_count = get_cart_count()

# ==========================================================
# TAB NAVIGATION
# ==========================================================
TAB_LABELS = [
    "📋 Browse & Add to Cart",
    f"🛒 Order Cart ({cart_count})",
    "📊 Stock Value",
    "🔎 Price Comparison"
]

active_tab = st.radio(
    "Navigation",
    TAB_LABELS,
    index=TAB_LABELS.index([t for t in TAB_LABELS if t.startswith(st.session_state.active_tab.split("(")[0])][0]),
    horizontal=True,
    label_visibility="collapsed",
    key="tab_radio",
)
st.session_state.active_tab = active_tab.split("(")[0].strip()
st.markdown("---")

# ══════════════════════════════════════════════════════════
# TAB 1 — BROWSE & ADD TO CART
# ══════════════════════════════════════════════════════════
if active_tab.startswith("📋"):

    st.markdown("## 📋 Browse Items to Order")
    if not reorder_available or df_unordered.empty:
        st.info("No unordered items found in the re_order table.")
    else:
        parsed_un   = df_unordered["GROUP"].apply(lambda g: (get_category(g), get_suppliers(g)))
        all_cats_un = sorted(set(c for c, _ in parsed_un if c))

        fc, fs, fbtn = st.columns([2.5, 2.5, 1.2])
        with fc:
            sel_cat = st.selectbox(
                "Filter by Category",
                ["— All Categories —"] + all_cats_un,
                key=f"t1_cat_{st.session_state.order_clear}"
            )

        all_sups_un = sorted(set(
            s for (c, sups) in parsed_un for s in sups
            if (sel_cat == "— All Categories —" or c == sel_cat)
        ))

        with fs:
            sel_sup = st.selectbox(
                "Filter by Supplier",
                ["— All Suppliers —"] + all_sups_un,
                key=f"t1_sup_{st.session_state.order_clear}"
            )

        with fbtn:
            st.markdown("<div style='padding-top:28px'>", unsafe_allow_html=True)
            if st.button("🔄 Clear Filters", type="secondary", use_container_width=True, key="t1_clear"):
                st.session_state.order_clear += 1
                st.session_state.draft_order_qty = {}
                st.rerun()
            st.markdown("</div>", unsafe_allow_html=True)

        disp = df_unordered.copy()
        if sel_cat != "— All Categories —":
            disp = disp[disp["GROUP"].apply(get_category) == sel_cat]
        if sel_sup != "— All Suppliers —":
            disp = disp[disp["GROUP"].apply(lambda g: sel_sup in get_suppliers(g))]

        for col in ["STOCK", "USAGE"]:
            disp[col] = pd.to_numeric(disp[col], errors="coerce").fillna(0)
        for col in ["COST PRICE", "SELLING PRICE"]:
            disp[col] = pd.to_numeric(disp[col], errors="coerce")

        disp = disp.sort_values("USAGE", ascending=False).reset_index(drop=True)

        # Add CATEGORY and restore ORDER QTY from session state
        disp["CATEGORY"] = disp["GROUP"].apply(get_category)
        disp["ORDER QTY"] = disp["PLU CODE"].astype(str).map(
            lambda x: st.session_state.draft_order_qty.get(x, None)
        )

        if sel_sup != "— All Suppliers —":
            mu, mv, _ = st.columns([1.8, 1.8, 5])
            mu.metric("📦 Units on Hand", f"{disp['STOCK'].sum():,.0f}")
            mv.metric("💲 Stock Value", f"${(disp['STOCK'] * disp['COST PRICE'].fillna(0)).sum():,.2f}")

        st.info(f"Found **{len(disp)}** items — enter quantities and add to cart")

        col_cfg = {
            "PLU CODE":      st.column_config.TextColumn(disabled=True),
            "DESCRIPTION":   st.column_config.TextColumn(disabled=True),
            "COST PRICE":    st.column_config.NumberColumn(disabled=True, format="$%.2f"),
            "SELLING PRICE": st.column_config.NumberColumn(disabled=True, format="$%.2f"),
            "GROUP":         st.column_config.TextColumn(disabled=True),
            "STOCK":         st.column_config.NumberColumn(disabled=True, format="%d"),
            "USAGE":         st.column_config.NumberColumn(disabled=True, format="%d"),
            "ORDER QTY":     st.column_config.NumberColumn(
                disabled=False, min_value=0, step=1, format="%d",
                help="Enter cases/units to order"
            ),
        }

        show_cols = ["PLU CODE","DESCRIPTION","COST PRICE","SELLING PRICE","GROUP","STOCK","USAGE","ORDER QTY"]
        edited = st.data_editor(
            disp[show_cols],
            column_config=col_cfg,
            hide_index=False,
            use_container_width=True,
            height=480,
            key=f"t1_editor_{st.session_state.order_clear}"
        )

        # Persist current editor values into session state
        current_visible_plus = set(edited["PLU CODE"].astype(str).tolist())
        for _, row in edited.iterrows():
            plu = str(row["PLU CODE"])
            qty = pd.to_numeric(row["ORDER QTY"], errors="coerce")

            if pd.notna(qty) and qty > 0:
                st.session_state.draft_order_qty[plu] = int(qty)
            else:
                st.session_state.draft_order_qty.pop(plu, None)

        qty_rows = edited[
            edited["ORDER QTY"].notna() &
            (pd.to_numeric(edited["ORDER QTY"], errors="coerce").fillna(0) > 0)
        ].copy()
        n = len(qty_rows)

        st.markdown("---")
        ac1, ac2, ac3 = st.columns([2, 2, 6])

        with ac1:
            if n > 0:
                st.success(f"✅ {n} item(s) ready to add")
            else:
                st.info("Enter quantities above")

        with ac2:
            if n > 0:
                if st.button("🛒 Add to Cart", type="primary", use_container_width=True):
                    cart_items = qty_rows.merge(
                        disp[["PLU CODE", "SUPPLIER", "CATEGORY"]],
                        on="PLU CODE",
                        how="left"
                    )
                    added_count = add_to_cart(cart_items)
                    st.success(f"✅ Added {added_count} items to cart!")

                    # Clear only the items that were added
                    for plu in cart_items["PLU CODE"].astype(str).tolist():
                        st.session_state.draft_order_qty.pop(plu, None)

                    st.session_state.order_clear += 1
                    st.rerun()

        with ac3:
            if n > 0:
                st.caption("💡 Tip: Items will be added/updated in your cart. Review in the Cart tab.")

    st.markdown("---")
    st.markdown("## 🔍 Product Search")
    sc, bc = st.columns([6, 1])

    with sc:
        q = st.text_input(
            "Search",
            placeholder="e.g. cumin OR 12345 (last 5 digits of barcode)",
            label_visibility="collapsed",
            value=st.session_state.last_search,
            key=f"sq_{st.session_state.search_clear}"
        )

    with bc:
        if st.button("🔄 Clear", type="secondary", use_container_width=True, key="t3_clear"):
            st.session_state.search_clear += 1
            st.session_state.last_search = ""
            st.rerun()

    if q:
        st.session_state.last_search = q

    if not q or len(q.strip()) < 3:
        st.info("Enter at least 3 characters to search.")
    else:
        q = q.strip()
        df_search["_b5"] = df_search["BARCODE"].astype(str).str[-5:]
        res = df_search[
            df_search["Description"].str.lower().str.contains(q.lower(), na=False) |
            df_search["_b5"].str.contains(q, na=False)
        ].drop(columns=["_b5"])

        if res.empty:
            st.warning("No matching products found.")
        else:
            st.markdown(f"### Results for **'{q}'**")
            f1, f2, f3, f4 = st.columns(4)

            desc_f = f1.text_input("Filter description", key=f"df_{st.session_state.search_clear}")
            if desc_f:
                res = res[res["Description"].str.lower().str.contains(desc_f.lower(), na=False)]

            if "Size" in res.columns:
                sz = f2.multiselect("Filter Size", sorted(res["Size"].dropna().unique()), key=f"szf_{st.session_state.search_clear}")
                if sz:
                    res = res[res["Size"].isin(sz)]

            if "SUPPLIER" in res.columns:
                sup = f3.multiselect("Filter Supplier", sorted(res["SUPPLIER"].dropna().unique()), key=f"spf_{st.session_state.search_clear}")
                if sup:
                    res = res[res["SUPPLIER"].isin(sup)]

            if not res.empty:
                vp        = res["Pc. Cost"].dropna() if "Pc. Cost" in res.columns else res["Price"].dropna()
                price_col = "Pc. Cost" if "Pc. Cost" in res.columns else "Price"
                if not vp.empty and vp.min() != vp.max():
                    pr = f4.slider(
                        "Pc. Cost Range",
                        float(vp.min()), float(vp.max()),
                        (float(vp.min()), float(vp.max())),
                        key=f"prf_{st.session_state.search_clear}"
                    )
                    res = res[(res[price_col] >= pr[0]) & (res[price_col] <= pr[1])]

            if res.empty:
                st.warning("No items match your filters.")
            else:
                sort_col = "Pc. Cost" if "Pc. Cost" in res.columns else "Price"
                show = ["BARCODE","ITEM NUM","Description","Size","Pack","Price",
                        "Pc. Cost","Sell Price","SUPPLIER","AISLE","STOCK","USAGE"]
                show = [c for c in show if c in res.columns]
                final = res[show].sort_values(["BARCODE", sort_col]).reset_index(drop=True)
                final.index += 1

                deduped = final.drop_duplicates(subset=["BARCODE"], keep="first") if "BARCODE" in final.columns else final

                st.markdown("---")
                ma, mb, mc, md, me = st.columns(5)
                ma.metric("Total Items", final["BARCODE"].nunique() if "BARCODE" in final.columns else len(final))
                mb.metric("Suppliers", final["SUPPLIER"].nunique() if "SUPPLIER" in final.columns else "N/A")
                if "Pc. Cost" in final.columns and not final["Pc. Cost"].dropna().empty:
                    mc.metric("Lowest Price", f"${final['Pc. Cost'].min():,.3f}")
                if "STOCK" in deduped.columns:
                    ts = pd.to_numeric(deduped["STOCK"], errors="coerce").fillna(0).sum()
                    md.metric("Total Stock", f"{ts:,.0f}")
                if "USAGE" in deduped.columns:
                    tu = pd.to_numeric(deduped["USAGE"], errors="coerce").fillna(0).sum()
                    me.metric("Total Usage", f"{tu:,.0f}")

                st.dataframe(final, hide_index=False, height=600, use_container_width=True)
                st.download_button(
                    "📥 Download Results",
                    data=final.to_csv(index=True),
                    file_name=f"{q}_results.csv",
                    mime="text/csv"
                )

# ══════════════════════════════════════════════════════════
# TAB 2 — ORDER CART
# ══════════════════════════════════════════════════════════
elif active_tab.startswith("🛒"):

    st.markdown("## 🛒 Your Order Cart")

    cart_df = get_cart()

    if cart_df.empty:
        st.info("🛒 Your cart is empty. Go to 'Browse & Add to Cart' tab to add items.")
    else:
        total_items = len(cart_df)
        total_qty = cart_df["ORDER QTY"].sum()
        total_value = (cart_df["COST PRICE"] * cart_df["ORDER QTY"]).sum()
        unique_suppliers = cart_df["SUPPLIER"].nunique()

        m1, m2, m3, m4, _ = st.columns([1.5, 1.5, 1.5, 1.5, 4])
        m1.metric("📦 Items", f"{total_items}")
        m2.metric("🔢 Total Qty", f"{total_qty:,.0f}")
        m3.metric("💲 Total Cost", f"${total_value:,.2f}")
        m4.metric("🏭 Suppliers", f"{unique_suppliers}")

        st.markdown("---")

        grouped = cart_df.groupby("SUPPLIER")

        for supplier, supplier_df in grouped:
            with st.expander(f"**🏭 {supplier}** — {len(supplier_df)} items | {supplier_df['ORDER QTY'].sum():,.0f} units", expanded=True):

                sc1, sc2, sc3, _ = st.columns([2, 2, 2, 4])
                sc1.metric("Items", len(supplier_df))
                sc2.metric("Order Qty", f"{supplier_df['ORDER QTY'].sum():,.0f}")
                sc3.metric("Total Cost", f"${(supplier_df['COST PRICE'] * supplier_df['ORDER QTY']).sum():,.2f}")

                col_cfg = {
                    "PLU CODE":      st.column_config.TextColumn(disabled=True),
                    "DESCRIPTION":   st.column_config.TextColumn(disabled=True),
                    "COST PRICE":    st.column_config.NumberColumn(disabled=True, format="$%.2f"),
                    "SELLING PRICE": st.column_config.NumberColumn(disabled=True, format="$%.2f"),
                    "STOCK":         st.column_config.NumberColumn(disabled=True, format="%d"),
                    "USAGE":         st.column_config.NumberColumn(disabled=True, format="%d"),
                    "ORDER QTY":     st.column_config.NumberColumn(
                        disabled=False, min_value=0, step=1, format="%d",
                        help="Edit quantity or set to 0 to remove"
                    ),
                }

                show_cols = ["PLU CODE", "DESCRIPTION", "COST PRICE", "SELLING PRICE", "STOCK", "USAGE", "ORDER QTY"]
                edited_supplier = st.data_editor(
                    supplier_df[show_cols].reset_index(drop=True),
                    column_config=col_cfg,
                    hide_index=False,
                    use_container_width=True,
                    height=min(300, len(supplier_df) * 35 + 50),
                    key=f"cart_editor_{supplier}_{st.session_state.cart_refresh}"
                )

                b1, b2, b3, _ = st.columns([2, 2, 2, 4])

                with b1:
                    if st.button(f"💾 Update {supplier}", type="secondary", use_container_width=True, key=f"update_{supplier}"):
                        for idx, row in edited_supplier.iterrows():
                            plu = row["PLU CODE"]
                            new_qty = row["ORDER QTY"]

                            if pd.isna(new_qty) or new_qty == 0:
                                delete_cart_item(plu)
                            else:
                                client = get_supabase_client()
                                client.table("cart_items").update({
                                    "order_qty": int(new_qty)
                                }).eq("plu_code", plu).execute()

                        st.success(f"✅ Updated {supplier} items!")
                        st.session_state.cart_refresh += 1
                        st.rerun()

                with b2:
                    excel_data = build_order_excel(
                        supplier_df[["PLU CODE", "DESCRIPTION", "COST PRICE", "SELLING PRICE", "GROUP", "STOCK", "USAGE", "ORDER QTY"]],
                        filename_prefix=supplier.replace(" ", "_")
                    )
                    st.download_button(
                        f"📥 Download {supplier}",
                        data=excel_data,
                        file_name=f"order_{supplier.replace(' ', '_')}.xlsx",
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                        use_container_width=True,
                        key=f"download_{supplier}"
                    )

                with b3:
                    if st.button(f"🗑️ Clear {supplier}", type="secondary", use_container_width=True, key=f"clear_{supplier}"):
                        clear_supplier_cart(supplier)
                        st.success(f"🗑️ Cleared {supplier} from cart!")
                        st.session_state.cart_refresh += 1
                        st.rerun()

        st.markdown("---")
        st.markdown("### 🎯 Cart Actions")

        ga1, ga2, _ = st.columns([2, 2, 6])

        with ga1:
            all_excel = build_order_excel(
                cart_df[["PLU CODE", "DESCRIPTION", "COST PRICE", "SELLING PRICE", "GROUP", "STOCK", "USAGE", "ORDER QTY"]],
                filename_prefix="all_suppliers"
            )
            st.download_button(
                "📥 Download All Orders",
                data=all_excel,
                file_name="order_all_suppliers.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                type="primary",
                use_container_width=True
            )

        with ga2:
            if st.button("🗑️ Clear Entire Cart", type="secondary", use_container_width=True):
                clear_cart()
                st.success("🗑️ Cart cleared!")
                st.session_state.cart_refresh += 1
                st.rerun()

# ══════════════════════════════════════════════════════════
# TAB 3 — STOCK VALUE
# ══════════════════════════════════════════════════════════
elif active_tab.startswith("📊"):

    st.markdown("## 📊 Stock Value — Current Inventory (Yellow PLU Codes)")
    if df_yfull.empty:
        st.info("No yellow PLU data found in the re_order table.")
    else:
        sv = df_yfull.copy()
        sv["STOCK"]       = pd.to_numeric(sv["STOCK"], errors="coerce").fillna(0)
        sv["COST"]        = pd.to_numeric(sv["COST"],  errors="coerce").fillna(0)
        sv["STOCK VALUE"] = sv["STOCK"] * sv["COST"]
        sv["CATEGORY"]    = sv["GROUP"].apply(get_category)

        def resolve_supplier_tab2(row):
            g2 = str(row.get("GROUP2", "")).strip()
            if g2 and g2 != "0" and g2.lower() != "none":
                return g2
            return ", ".join(get_suppliers(row["GROUP"]))

        sv["SUPPLIER"] = sv.apply(resolve_supplier_tab2, axis=1)

        ma, mb, mc, md = st.columns(4)
        ma.metric("📦 Total SKUs", f"{len(sv):,}")
        mb.metric("🔢 Total Units", f"{sv['STOCK'].sum():,.0f}")
        mc.metric("💲 Total Stock Value", f"${sv['STOCK VALUE'].sum():,.2f}")
        md.metric("✅ SKUs with Stock", f"{len(sv[sv['STOCK'] > 0]):,}")

        st.markdown("---")

        all_cats_sv = sorted([c for c in sv["CATEGORY"].dropna().unique() if c])
        fmode, fc2, fs2, fbtn2 = st.columns([2, 2.5, 2.5, 0.8])

        with fmode:
            st.markdown("<div style='padding-top:4px; font-size:14px; color:#555'>View grouped by</div>", unsafe_allow_html=True)
            rc1, rc2 = st.columns(2)
            with rc1:
                if st.button(
                    "📂 Category",
                    type="primary" if st.session_state.get("sv_mode","Category") == "Category" else "secondary",
                    use_container_width=True,
                    key="sv_mode_cat"
                ):
                    st.session_state.sv_mode = "Category"
                    st.rerun()
            with rc2:
                if st.button(
                    "🏭 Supplier",
                    type="primary" if st.session_state.get("sv_mode","Category") == "Supplier" else "secondary",
                    use_container_width=True,
                    key="sv_mode_sup"
                ):
                    st.session_state.sv_mode = "Supplier"
                    st.rerun()
            view_mode = st.session_state.get("sv_mode", "Category")

        with fc2:
            sel_cat2 = st.selectbox("Filter by Category", ["— All Categories —"] + all_cats_sv, key=f"sv_cat_{st.session_state.sv_clear}")

        pool = sv if sel_cat2 == "— All Categories —" else sv[sv["CATEGORY"] == sel_cat2]
        all_sups_sv = sorted(set(s for s in pool["SUPPLIER"] if s and s.strip()))

        with fs2:
            sel_sup2 = st.selectbox("Filter by Supplier", ["— All Suppliers —"] + all_sups_sv, key=f"sv_sup_{st.session_state.sv_clear}")

        with fbtn2:
            st.markdown("<div style='padding-top:28px'>", unsafe_allow_html=True)
            if st.button("🔄 Clear", type="secondary", use_container_width=True, key="sv_clear_btn"):
                st.session_state.sv_clear += 1
                st.rerun()
            st.markdown("</div>", unsafe_allow_html=True)

        filt = sv.copy()
        if sel_cat2 != "— All Categories —":
            filt = filt[filt["CATEGORY"] == sel_cat2]
        if sel_sup2 != "— All Suppliers —":
            filt = filt[filt["SUPPLIER"] == sel_sup2]

        if sel_cat2 != "— All Categories —" or sel_sup2 != "— All Suppliers —":
            label_parts = []
            if sel_cat2 != "— All Categories —":
                label_parts.append(sel_cat2)
            if sel_sup2 != "— All Suppliers —":
                label_parts.append(sel_sup2)

            st.markdown(f"### 📌 Showing: **{' / '.join(label_parts)}**")
            r1, r2, r3, r4 = st.columns(4)
            r1.metric("📦 SKUs", f"{len(filt):,}")
            r2.metric("🔢 Units on Hand", f"{filt['STOCK'].sum():,.0f}")
            r3.metric("💲 Stock Value", f"${filt['STOCK VALUE'].sum():,.2f}")
            r4.metric("✅ SKUs with Stock", f"{len(filt[filt['STOCK'] > 0]):,}")
            st.markdown("---")

        st.markdown(f"### {'📂 By Category' if view_mode == 'Category' else '🏭 By Supplier'}")

        if view_mode == "Category":
            grp = (
                filt.groupby("CATEGORY")
                .agg(SKUs=("PLU CODE", "count"), Units=("STOCK", "sum"), Stock_Value=("STOCK VALUE", "sum"))
                .reset_index()
                .rename(columns={"CATEGORY": "Category", "Stock_Value": "Stock Value ($)"})
                .sort_values("Stock Value ($)", ascending=False)
            )
            grp["Stock Value ($)"] = grp["Stock Value ($)"].map("${:,.2f}".format)
            grp["Units"] = grp["Units"].map("{:,.0f}".format)
            st.dataframe(grp, use_container_width=True, hide_index=True, height=420)
        else:
            grp = (
                filt.groupby("SUPPLIER")
                .agg(SKUs=("PLU CODE", "count"), Units=("STOCK", "sum"), Stock_Value=("STOCK VALUE", "sum"))
                .reset_index()
                .rename(columns={"SUPPLIER": "Supplier", "Stock_Value": "Stock Value ($)"})
                .sort_values("Stock Value ($)", ascending=False)
            )
            grp["Stock Value ($)"] = grp["Stock Value ($)"].map("${:,.2f}".format)
            grp["Units"] = grp["Units"].map("{:,.0f}".format)
            st.dataframe(grp, use_container_width=True, hide_index=True, height=420)

        st.markdown("---")
        with st.expander("🔍 View individual items", expanded=False):
            detail = filt[["PLU CODE","DESCRIPTION","CATEGORY","SUPPLIER","COST","STOCK","STOCK VALUE"]].copy()
            detail = detail.sort_values("STOCK VALUE", ascending=False).reset_index(drop=True)
            detail.index += 1
            detail["COST"] = detail["COST"].map("${:,.2f}".format)
            detail["STOCK VALUE"] = detail["STOCK VALUE"].map("${:,.2f}".format)
            st.dataframe(detail, use_container_width=True, height=420)

            dl = filt[["PLU CODE","DESCRIPTION","CATEGORY","SUPPLIER","COST","STOCK","STOCK VALUE"]].copy()
            st.download_button(
                "📥 Download Stock Value Report (.csv)",
                data=dl.to_csv(index=False),
                file_name="stock_value_report.csv",
                mime="text/csv"
            )

# ══════════════════════════════════════════════════════════
# TAB 4 — PRICE COMPARISON
# ══════════════════════════════════════════════════════════
elif active_tab.startswith("🔎"):

    st.markdown("## 🔎 Price Comparison")

    if df_yfull.empty:
        st.info("No data found in the re_order table.")
    else:
        TOL = 0.01

        def match_status(a, b):
            if pd.isna(a) or pd.isna(b):
                return "⚠️ Missing"
            return "✅ Match" if abs(a - b) <= TOL else "❌ Mismatch"

        comp = df_yfull[["PLU CODE","DESCRIPTION","COST"]].copy()
        comp = comp.merge(df_price1, on="PLU CODE", how="left")

        ep_cols  = ["BARCODE","Price","Pc. Cost","Sell Price","Description","SUPPLIER"]
        ep_avail = [c for c in ep_cols if c in df_prices.columns]
        ep = df_prices[ep_avail].drop_duplicates(subset=["BARCODE"])

        comp = comp.merge(ep, left_on="PLU CODE", right_on="BARCODE", how="left")
        comp = comp.drop(columns=["BARCODE"], errors="ignore")

        for col in ["COST","PRICE 1","Pc. Cost","Sell Price","Price"]:
            if col in comp.columns:
                comp[col] = pd.to_numeric(comp[col], errors="coerce")

        comp["COST MATCH"]    = comp.apply(lambda r: match_status(r["COST"], r.get("Pc. Cost")), axis=1)
        comp["SELLING MATCH"] = comp.apply(lambda r: match_status(r.get("PRICE 1"), r.get("Sell Price")), axis=1)
        comp = comp.reset_index(drop=True)

        all_cost = comp["COST MATCH"].value_counts()
        all_sell = comp["SELLING MATCH"].value_counts()

        total_mismatches = len(comp[
            (comp["COST MATCH"] == "❌ Mismatch") |
            (comp["SELLING MATCH"] == "❌ Mismatch")
        ])

        if total_mismatches == 0:
            st.success("✅ All RE ORDER prices match EXISTING PRICES.")
        else:
            st.warning(
                f"⚠️ **{total_mismatches}** item(s) have mismatched prices "
                f"between RE ORDER and EXISTING PRICES."
            )

        st.markdown("#### 💰 Cost Price Comparison")
        st.caption("RE ORDER sheet (COST col) vs EXISTING PRICES sheet (Pc. Cost col)")

        cost_filter = st.selectbox(
            "Filter by match status",
            ["All","✅ Match","❌ Mismatch","⚠️ Missing"],
            key="cost_filter"
        )

        cost_df = comp[["PLU CODE","DESCRIPTION","COST","Pc. Cost","COST MATCH"]].copy()
        cost_df.columns = ["PLU CODE","DESCRIPTION","RE ORDER Cost","Existing Pc. Cost","Status"]

        if cost_filter != "All":
            cost_df = cost_df[cost_df["Status"] == cost_filter]

        cc1, cc2, cc3, cc4 = st.columns(4)
        cc1.metric("Total", len(comp))
        cc2.metric("✅ Match", int(all_cost.get("✅ Match", 0)))
        cc3.metric("❌ Mismatch", int(all_cost.get("❌ Mismatch", 0)))
        cc4.metric("⚠️ Missing", int(all_cost.get("⚠️ Missing", 0)))

        st.dataframe(
            cost_df.reset_index(drop=True),
            use_container_width=True,
            height=380,
            column_config={
                "RE ORDER Cost":     st.column_config.NumberColumn(format="$%.2f"),
                "Existing Pc. Cost": st.column_config.NumberColumn(format="$%.2f"),
                "Status":            st.column_config.TextColumn("Status"),
            },
            hide_index=True
        )

        st.markdown("---")

        st.markdown("#### 🏷️ Selling Price Comparison")
        st.caption("RE ORDER sheet (PRICE 1 col) vs EXISTING PRICES sheet (Sell Price col)")

        sell_filter = st.selectbox(
            "Filter by match status",
            ["All","✅ Match","❌ Mismatch","⚠️ Missing"],
            key="sell_filter"
        )

        sell_df = comp[["PLU CODE","DESCRIPTION","PRICE 1","Sell Price","SELLING MATCH"]].copy()
        sell_df.columns = ["PLU CODE","DESCRIPTION","RE ORDER Price 1","Existing Sell Price","Status"]

        if sell_filter != "All":
            sell_df = sell_df[sell_df["Status"] == sell_filter]

        sc1, sc2, sc3, sc4 = st.columns(4)
        sc1.metric("Total", len(comp))
        sc2.metric("✅ Match", int(all_sell.get("✅ Match", 0)))
        sc3.metric("❌ Mismatch", int(all_sell.get("❌ Mismatch", 0)))
        sc4.metric("⚠️ Missing", int(all_sell.get("⚠️ Missing", 0)))

        st.dataframe(
            sell_df.reset_index(drop=True),
            use_container_width=True,
            height=380,
            column_config={
                "RE ORDER Price 1":    st.column_config.NumberColumn(format="$%.2f"),
                "Existing Sell Price": st.column_config.NumberColumn(format="$%.2f"),
                "Status":              st.column_config.TextColumn("Status"),
            },
            hide_index=True
        )
